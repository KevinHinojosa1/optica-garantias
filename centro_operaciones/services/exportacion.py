"""Exportación Excel con formato para matrices de seguimiento."""

from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _estilo_encabezado(ws, fila: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for cell in ws[fila]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def _auto_ancho(ws, max_width: int = 45) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        largo = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(largo + 2, 10), max_width)


def exportar_matriz_seguimiento(df: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    export = df.copy()
    if "fecha_alerta" in export.columns:
        export["fecha_alerta"] = pd.to_datetime(export["fecha_alerta"], errors="coerce").dt.strftime("%d/%m/%Y %H:%M")

    resumen = (
        export.groupby(["local", "clasificacion"], dropna=False)
        .size()
        .reset_index(name="casos")
        .sort_values("casos", ascending=False)
    )
    pendientes = export[export["estado_gestion"] == "Sin gestión"]

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        export.to_excel(writer, index=False, sheet_name="Matriz Seguimiento")
        resumen.to_excel(writer, index=False, sheet_name="Resumen Local")
        pendientes.to_excel(writer, index=False, sheet_name="Pendientes")

    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb["Matriz Seguimiento"]
    _estilo_encabezado(ws)
    ws.freeze_panes = "A2"
    fill_pend = PatternFill("solid", fgColor="FFF3CD")
    fill_res = PatternFill("solid", fgColor="D1E7DD")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        estado = str(row[13].value or "") if len(row) > 13 else ""
        if estado == "Sin gestión":
            for cell in row:
                cell.fill = fill_pend
        elif estado == "Resuelto":
            for cell in row:
                cell.fill = fill_res
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _auto_ancho(ws)
    for nombre in ("Resumen Local", "Pendientes"):
        if nombre in wb.sheetnames:
            _estilo_encabezado(wb[nombre])
            _auto_ancho(wb[nombre])

    meta = wb.create_sheet("_Meta", 0)
    meta["A1"] = "Centro de Operaciones — Óptica Los Andes"
    meta["A2"] = f"Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    meta["A3"] = f"Total casos: {len(export)}"
    meta["A4"] = f"Pendientes sin gestión: {len(pendientes)}"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()